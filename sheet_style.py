from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os


FONT_BOLD = Font(bold=True)
class SheetStyle:


    def __init__(self):
        
        file_path = os.getcwd() + r"\Carteira.xlsx"
        self.wb = load_workbook(file_path)
        self.ws_wallet = self.wb["Carteira"]
        self.ws_dividends = self.wb["Proventos"]


    def wallet_style(self):
        last_row = self.ws_wallet.max_row + 2
        last_column_number = self.ws_wallet.max_column
        last_column = get_column_letter(last_column_number)

        self.ws_wallet[last_column+str(last_row)] = f"=SUM({last_column}2:{last_column}{last_row-1})"
        self.ws_wallet[last_column + str(last_row)].font = FONT_BOLD

        self.ws_wallet["F" + str(last_row)] = f"=SUM(F2:F{last_row-1})"
        self.ws_wallet["F" + str(last_row)].font = FONT_BOLD

        self.ws_wallet["G" + str(last_row)] = f"=SUM(G2:G{last_row-1})"
        self.ws_wallet["G" + str(last_row)].font = FONT_BOLD

        for row in self.ws_wallet.iter_rows(min_row=2, min_col=4, max_col=4, max_row=last_row):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        for row in self.ws_wallet.iter_rows(min_row=2, min_col=6, max_col=last_column_number,
                                            max_row=last_row):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        total_dividends_column = get_column_letter(last_column_number - 1)
        self.ws_wallet[total_dividends_column + str(last_row)] = "TOTAL DE PROVENTOS =>"
        self.ws_wallet[total_dividends_column + str(last_row)].font = FONT_BOLD

    def dividends_style(self):
        self.ws_dividends = self.wb["Proventos"]

        last_row = self.ws_dividends.max_row
        last_column_number = self.ws_dividends.max_column
        value_column = get_column_letter(last_column_number)
        sum_column = get_column_letter(last_column_number + 1)

        self.ws_dividends[sum_column + "2"] = f"=SUM({value_column}2:{value_column}{last_row})"
        self.ws_dividends[sum_column + "2"].number_format = 'R$ #,##0.00'
        for row in self.ws_dividends.iter_rows(min_row=2, min_col=2, max_col=2, max_row=last_row):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'


    def auto_fit(self, sheet):

        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width
        return sheet


    def run(self):
        self.wallet_style()
        self.dividends_style()
        self.auto_fit(self.ws_wallet)
        self.auto_fit(self.ws_dividends)
        self.wb.save("Carteira.xlsx")